# rag_evaluate_trulens.py — TruLens RAG Triad evaluation for the hybrid RAG
#
# 5-step scientific method (mirrors rag_evaluate.py and rag_evaluate_deepeval.py):
#
#  1. OBSERVATION   — LLMs hallucinate; RAG is meant to fix that.
#  2. HYPOTHESIS    — TruLens RAG Triad (Context Relevance, Groundedness,
#                     Answer Relevance) will verify where the system succeeds
#                     and fails, with chain-of-thought reasoning per score.
#  3. EXPERIMENT    — Run the full hybrid-RAG pipeline against the same
#                     GROUND_TRUTH dataset used in RAGAS and DeepEval.
#  4. ANALYSIS      — Score each question, print a leaderboard table,
#                     export a colour-coded Excel report.
#  5. CONCLUSION    — Compare TruLens scores with RAGAS and DeepEval to get
#                     a three-framework view of RAG quality.
#
# TruLens RAG Triad — three feedback functions (all higher = better):
#
#  • Context Relevance  — Is each retrieved chunk relevant to the query?
#                         (maps to RAGAS Context Precision / DeepEval ContextualPrecision)
#  • Groundedness       — Is the answer grounded in the retrieved context?
#                         (maps to RAGAS Faithfulness / DeepEval Faithfulness)
#  • Answer Relevance   — Does the answer address the original question?
#                         (maps to RAGAS Answer Relevancy / DeepEval AnswerRelevancy)
#
# KEY DIFFERENCE vs RAGAS and DeepEval:
#  TruLens uses chain-of-thought (CoT) reasoning for every score, meaning the
#  judge LLM explains its reasoning step-by-step before assigning a number.
#  This makes scores more interpretable and easier to debug than RAGAS
#  (free-text parse) or DeepEval (structured JSON).
#
# Install:
#   pip install trulens trulens-providers-litellm litellm \
#               chromadb pypdf nltk rank_bm25 openpyxl requests
#
# Verified against: trulens==2.8.1, trulens-providers-litellm==2.8.1
#
# Imports confirmed working in trulens 2.8.1:
#   from trulens.core import TruSession, Feedback, Select
#   from trulens.apps.app import instrument          ← correct path in 2.8.1
#   from trulens.apps.custom import TruCustomApp
#   from trulens.providers.litellm import LiteLLM
#
# Run:
#   python rag_evaluate_trulens.py
#
# Output:
#   trulens_report_<timestamp>.xlsx
#
# NOTE on Ollama API endpoint:
#   Ollama >= 0.2 uses /api/embed (not /api/embeddings):
#     request:  {"model": ..., "input": text}
#     response: {"embeddings": [[...float...]]}

import os
import time
import hashlib
import requests
import numpy as np
from pathlib import Path
from datetime import datetime

import nltk
import chromadb
from rank_bm25 import BM25Okapi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── TruLens v2.8.1 imports ───────────────────────────────────────────────────
# trulens.apps.app is the correct path for `instrument` in 2.8.1
# (trulens.apps.custom still works but emits a DeprecationWarning)
from trulens.core import TruSession, Feedback, Select
from trulens.apps.app import instrument
from trulens.apps.custom import TruCustomApp
from trulens.providers.litellm import LiteLLM

import litellm
litellm.set_verbose = False   # suppress LiteLLM noise

nltk.download("punkt",     quiet=True)
nltk.download("punkt_tab", quiet=True)

# ─── CONFIG ──────────────────────────────────────────────────────────────────

OLLAMA_URL      = os.getenv("OLLAMA_URL", "http://localhost:11434")
GEN_MODEL       = "llama3.2"
EMBED_MODEL     = os.getenv("EMBED_MODEL", "mxbai-embed-large")
DOCS_DIR        = "./docs"
CHROMA_DIR      = "./chroma_store"
COLLECTION      = "rag_knowledge"
CHUNK_SIZE      = 1000
CHUNK_OVERLAP   = 150
TOP_K           = 5
SCORE_THRESHOLD = 0.40
RRF_K           = 60

# ── TruLens LiteLLM provider (Ollama, no OpenAI key) ─────────────────────────
# LiteLLM bridges TruLens to your local Ollama instance.
# model_engine format: "ollama/<model_name>"
provider = LiteLLM(
    model_engine = f"ollama/{GEN_MODEL}",
    api_base     = OLLAMA_URL,
)

# ─── GROUND TRUTH DATASET ────────────────────────────────────────────────────
# Identical to rag_evaluate.py and rag_evaluate_deepeval.py
# so scores are directly comparable across all three frameworks.

GROUND_TRUTH = [
    # ── Category A: Factual / Definition ────────────────────────────────────
    {
        "question": "What is Tier 1 capital?",
        "expected_output": (
            "Tier 1 capital is the going-concern capital of a bank. "
            "It consists of Common Equity Tier 1 (CET1) capital and "
            "Additional Tier 1 (AT1) capital. CET1 is the highest quality "
            "capital and includes common shares, retained earnings, and "
            "other comprehensive income. AT1 instruments are subordinated, "
            "have no maturity, and can absorb losses on a going-concern basis."
        ),
        "ground_truth_pages": [101, 102],
    },
    {
        "question": "What is the definition of a Global Systemically Important Bank (G-SIB)?",
        "expected_output": (
            "A Global Systemically Important Bank (G-SIB) is a bank whose "
            "distress or failure would cause significant disruption to the "
            "broader global financial system and economic activity. G-SIBs "
            "are identified through an assessment methodology based on five "
            "categories: size, interconnectedness, substitutability, "
            "complexity, and cross-border activity."
        ),
        "ground_truth_pages": [17, 18],
    },
    {
        "question": "What does LCR stand for and what is its minimum requirement?",
        "expected_output": (
            "LCR stands for Liquidity Coverage Ratio. The minimum requirement "
            "is 100%, meaning a bank must hold sufficient high-quality liquid "
            "assets (HQLA) to cover total net cash outflows over a 30-day "
            "stressed period."
        ),
        "ground_truth_pages": [1093, 1097],
    },
    {
        "question": "What is the Net Stable Funding Ratio (NSFR)?",
        "expected_output": (
            "The Net Stable Funding Ratio (NSFR) requires banks to maintain "
            "a stable funding profile in relation to their assets and "
            "off-balance sheet activities over a one-year horizon. "
            "The NSFR is defined as the ratio of Available Stable Funding "
            "to Required Stable Funding, and must be at least 100%."
        ),
        "ground_truth_pages": [1216, 1219],
    },
    {
        "question": "How does the Basel Framework define a trading book?",
        "expected_output": (
            "The trading book consists of positions in financial instruments "
            "and commodities held either with trading intent or to hedge "
            "other elements of the trading book. Trading intent is evidenced "
            "by the strategies, policies, and procedures a bank has in place "
            "to manage the position."
        ),
        "ground_truth_pages": [201, 202],
    },
    # ── Category B: Numerical / Threshold ───────────────────────────────────
    {
        "question": "What is the minimum Common Equity Tier 1 (CET1) ratio under Basel III?",
        "expected_output": (
            "The minimum Common Equity Tier 1 (CET1) ratio under Basel III "
            "is 4.5% of risk-weighted assets."
        ),
        "ground_truth_pages": [193],
    },
    {
        "question": "What is the capital conservation buffer requirement?",
        "expected_output": (
            "The capital conservation buffer is set at 2.5% of total "
            "risk-weighted assets and must be met with Common Equity Tier 1 "
            "capital. When a bank's CET1 falls within the buffer range, "
            "constraints are placed on earnings distributions."
        ),
        "ground_truth_pages": [217, 218],
    },
    {
        "question": "What is the large exposure limit as a percentage of Tier 1 capital?",
        "expected_output": (
            "The large exposure limit is 25% of a bank's eligible Tier 1 "
            "capital. For exposures between G-SIBs, a stricter limit of "
            "15% of Tier 1 capital applies."
        ),
        "ground_truth_pages": [1260],
    },
    # ── Category C: Process / How-to ────────────────────────────────────────
    {
        "question": "How does a bank calculate its Liquidity Coverage Ratio?",
        "expected_output": (
            "The LCR is calculated as the stock of high-quality liquid assets "
            "(HQLA) divided by the total net cash outflows over the next "
            "30 calendar days under a stress scenario. The result must be "
            "at least 100%. Net cash outflows are calculated as total expected "
            "outflows minus total expected inflows, capped at 75% of total outflows."
        ),
        "ground_truth_pages": [1097, 1098],
    },
    {
        "question": "How is the leverage ratio calculated?",
        "expected_output": (
            "The leverage ratio is calculated as Tier 1 capital divided by "
            "the bank's total exposure measure. The exposure measure includes "
            "on-balance sheet assets, derivative exposures, securities "
            "financing transaction exposures, and off-balance sheet items. "
            "The minimum leverage ratio requirement is 3% of Tier 1 capital."
        ),
        "ground_truth_pages": [1052, 1054],
    },
    # ── Category D: Comparative ──────────────────────────────────────────────
    {
        "question": "What is the difference between the Standardised Approach and the IRB approach for credit risk?",
        "expected_output": (
            "Under the Standardised Approach, risk weights are assigned by "
            "supervisors based on external credit ratings or asset class. "
            "Under the Internal Ratings-Based (IRB) approach, banks use "
            "their own internal models to estimate risk components such as "
            "Probability of Default (PD), Loss Given Default (LGD), and "
            "Exposure at Default (EAD), subject to supervisory approval."
        ),
        "ground_truth_pages": [248, 343],
    },
    {
        "question": "What distinguishes Tier 1 capital from Tier 2 capital?",
        "expected_output": (
            "Tier 1 capital is going-concern capital — it absorbs losses "
            "while the bank continues to operate. Tier 2 capital is "
            "gone-concern capital — it absorbs losses only in liquidation. "
            "Tier 2 instruments include subordinated debt with a minimum "
            "maturity of five years and certain loan loss provisions."
        ),
        "ground_truth_pages": [101, 125],
    },
    # ── Category E: Out-of-scope ─────────────────────────────────────────────
    {
        "question": "What is the current ECB interest rate?",
        "expected_output": "I don't know. This information is not in the Basel Framework document.",
        "ground_truth_pages": [],
    },
    {
        "question": "What does IFRS 9 say about expected credit loss provisioning?",
        "expected_output": "I don't know. IFRS 9 details are not covered in the Basel Framework document.",
        "ground_truth_pages": [],
    },
    # ── Category F: Advanced / Multi-hop ────────────────────────────────────
    {
        "question": "How does a bank's CET1 ratio affect dividend eligibility under the capital conservation buffer?",
        "expected_output": (
            "When a bank's CET1 ratio falls within the capital conservation "
            "buffer range (4.5% to 7%), earnings distribution constraints "
            "are imposed. The closer the CET1 ratio is to the 4.5% minimum, "
            "the greater the restriction on dividends, share buybacks, and "
            "discretionary bonus payments. A bank with a CET1 ratio between "
            "4.5% and 5.125% must retain at least 100% of its earnings."
        ),
        "ground_truth_pages": [217, 218, 219],
    },
    {
        "question": "How are cryptoasset exposures classified and what capital treatment do they receive?",
        "expected_output": (
            "Cryptoassets are divided into two groups. Group 1 cryptoassets "
            "meet classification conditions and receive treatment broadly "
            "equivalent to traditional assets. Group 2 cryptoassets do not "
            "meet these conditions and are subject to a conservative capital "
            "treatment including a risk weight of 1250% for Group 2b assets."
        ),
        "ground_truth_pages": [37, 38],
    },
]

# ─── OLLAMA HELPERS ──────────────────────────────────────────────────────────
# /api/embed is correct for Ollama >= 0.2

def ollama_embed(texts: list) -> list:
    embeddings = []
    for text in texts:
        r = requests.post(
            f"{OLLAMA_URL}/api/embed",
            json={"model": EMBED_MODEL, "input": text},
        )
        r.raise_for_status()
        embeddings.append(r.json()["embeddings"][0])
    return embeddings


def ollama_generate(prompt: str) -> str:
    r = requests.post(
        f"{OLLAMA_URL}/api/generate",
        json={"model": GEN_MODEL, "prompt": prompt, "stream": False},
    )
    r.raise_for_status()
    return r.json()["response"].strip()


# ─── DOCUMENT LOADING ────────────────────────────────────────────────────────

def load_pdf(path: str) -> list:
    from pypdf import PdfReader
    reader = PdfReader(path)
    pages  = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append({"text": text, "page": i + 1})
    return pages


def load_documents(docs_dir: str) -> list:
    docs = []
    for filepath in Path(docs_dir).glob("**/*"):
        if filepath.suffix.lower() == ".pdf":
            pages = load_pdf(str(filepath))
            if pages:
                docs.append({"source": str(filepath), "pages": pages})
                print(f"  Loaded: {filepath.name} ({len(pages)} pages)")
        elif filepath.suffix.lower() == ".txt":
            with open(str(filepath), "r", encoding="utf-8") as f:
                text = f.read()
            if text.strip():
                docs.append({
                    "source": str(filepath),
                    "pages": [{"text": text, "page": None}],
                })
                print(f"  Loaded: {filepath.name}")
    return docs


# ─── CHUNKING ────────────────────────────────────────────────────────────────

def chunk_text(text: str, chunk_size: int, overlap: int) -> list:
    from nltk.tokenize import sent_tokenize
    sentences   = sent_tokenize(text)
    chunks      = []
    current     = []
    current_len = 0
    for sent in sentences:
        slen = len(sent)
        if current_len + slen > chunk_size and current:
            chunks.append(" ".join(current))
            kept, kept_len = [], 0
            for s in reversed(current):
                if kept_len + len(s) > overlap:
                    break
                kept.insert(0, s)
                kept_len += len(s)
            current, current_len = kept, kept_len
        current.append(sent)
        current_len += slen
    if current:
        chunks.append(" ".join(current))
    return [c for c in chunks if len(c) > 30]


def chunk_documents(docs: list) -> list:
    all_chunks = []
    for doc in docs:
        for pb in doc["pages"]:
            page_num = pb["page"] or 0
            for i, chunk in enumerate(chunk_text(pb["text"], CHUNK_SIZE, CHUNK_OVERLAP)):
                cid = hashlib.md5(
                    f"{doc['source']}_{page_num}_{i}".encode()
                ).hexdigest()
                all_chunks.append({
                    "id":          cid,
                    "text":        chunk,
                    "source":      doc["source"],
                    "page":        page_num,
                    "chunk_index": i,
                })
    return all_chunks


# ─── CHROMADB ────────────────────────────────────────────────────────────────

def get_collection():
    client = chromadb.PersistentClient(path=CHROMA_DIR)
    return client.get_or_create_collection(
        name=COLLECTION, metadata={"hnsw:space": "cosine"}
    )


def index_chunks(collection, chunks: list):
    existing = set(collection.get()["ids"])
    new      = [c for c in chunks if c["id"] not in existing]
    if not new:
        print("  ChromaDB: all chunks already indexed.")
        return
    print(f"  Embedding {len(new)} new chunks...")
    for i, chunk in enumerate(new):
        emb = ollama_embed([chunk["text"]])[0]
        collection.add(
            ids        = [chunk["id"]],
            embeddings = [emb],
            documents  = [chunk["text"]],
            metadatas  = [{
                "source":      chunk["source"],
                "page":        chunk["page"],
                "chunk_index": chunk["chunk_index"],
            }],
        )
        if (i + 1) % 100 == 0:
            print(f"  {i+1}/{len(new)}...")
    print("  Done.")


# ─── BM25 ────────────────────────────────────────────────────────────────────

class BM25Index:
    def __init__(self, chunks: list):
        self.chunks = chunks
        self.bm25   = BM25Okapi([c["text"].lower().split() for c in chunks])

    def search(self, query: str, top_k: int) -> list:
        scores = self.bm25.get_scores(query.lower().split())
        ranked = sorted(zip(scores, self.chunks), key=lambda x: x[0], reverse=True)[:top_k]
        return [
            {**chunk, "bm25_score": round(float(score), 4)}
            for score, chunk in ranked if score > 0
        ]


# ─── HYBRID RETRIEVAL ────────────────────────────────────────────────────────

def hybrid_retrieve(query: str, collection, bm25_index: BM25Index) -> list:
    qemb = ollama_embed([query])[0]
    res  = collection.query(
        query_embeddings = [qemb],
        n_results        = TOP_K,
        include          = ["documents", "metadatas", "distances"],
    )
    semantic = []
    for text, meta, dist in zip(
        res["documents"][0], res["metadatas"][0], res["distances"][0]
    ):
        score = round(1 - dist, 3)
        if score >= SCORE_THRESHOLD:
            semantic.append({
                "text":      text,
                "source":    Path(meta["source"]).name,
                "page":      meta.get("page", 0),
                "sem_score": score,
            })

    bm25_hits = [
        {
            "text":       c["text"],
            "source":     Path(c["source"]).name,
            "page":       c["page"],
            "bm25_score": c["bm25_score"],
        }
        for c in bm25_index.search(query, TOP_K)
    ]

    rrf_scores, chunk_store = {}, {}
    for rank, chunk in enumerate(semantic):
        key = chunk["text"][:60]
        rrf_scores[key]  = rrf_scores.get(key, 0) + 1 / (rank + RRF_K)
        chunk_store[key] = chunk
    for rank, chunk in enumerate(bm25_hits):
        key = chunk["text"][:60]
        rrf_scores[key]  = rrf_scores.get(key, 0) + 1 / (rank + RRF_K)
        if key not in chunk_store:
            chunk_store[key] = chunk

    ranked = sorted(rrf_scores.items(), key=lambda x: x[1], reverse=True)
    results = []
    for key, rrf_score in ranked[:TOP_K]:
        chunk = chunk_store[key].copy()
        chunk["rrf_score"] = round(rrf_score, 6)
        results.append(chunk)
    return results


# ─── GENERATION ──────────────────────────────────────────────────────────────

def generate_answer(question: str, retrieved: list) -> str:
    if not retrieved:
        return "I could not find relevant information in the indexed documents."
    context = "\n\n".join(
        f"[Source: {r['source']}, Page: {r['page']}]\n{r['text']}"
        for r in retrieved
    )
    prompt = f"""Use only the context below to answer the question.
Be concise and factual.
If the answer is not in the context, say "I don't know".
At the end of your answer, cite the source file and page number(s):
Source: <filename>, Page <number>

Context:
{context}

Question: {question}

Answer:"""
    return ollama_generate(prompt)


# ─── TruLens RAG TRIAD SCORING ───────────────────────────────────────────────
#
# We call provider methods directly (synchronous, batch-friendly).
# This avoids the complexity of TruCustomApp context managers for batch eval.
#
# Provider methods confirmed available in trulens-providers-litellm 2.8.1:
#   provider.context_relevance_with_cot_reasons(question, context_chunk)
#   provider.groundedness_measure_with_cot_reasons(full_context, answer)
#   provider.relevance_with_cot_reasons(question, answer)
#
# All return a tuple: (float_score, dict_with_reasons)
# Score is 0.0–1.0. All three metrics: higher = better.

def _call_provider(fn, *args):
    """Call a provider feedback function and safely unpack (score, reason)."""
    try:
        result = fn(*args)
        if isinstance(result, tuple) and len(result) == 2:
            score  = round(float(result[0]), 3)
            reason = result[1].get("reason", "") if isinstance(result[1], dict) else str(result[1])
        else:
            score  = round(float(result), 3)
            reason = ""
        return score, reason
    except Exception as e:
        return None, f"ERROR: {e}"


def run_trulens_metrics(
    question:  str,
    answer:    str,
    retrieved: list,
) -> tuple:
    context_texts = [r["text"] for r in retrieved]
    scores  = {}
    reasons = {}

    # 1. Context Relevance — evaluated per chunk, averaged
    ctx_scores, ctx_reasons = [], []
    for ctx in context_texts:
        s, r = _call_provider(
            provider.context_relevance_with_cot_reasons, question, ctx
        )
        ctx_scores.append(s)
        ctx_reasons.append(r)

    valid = [s for s in ctx_scores if s is not None]
    scores["context_relevance"]  = round(sum(valid) / len(valid), 3) if valid else None
    reasons["context_relevance"] = " | ".join(r for r in ctx_reasons if r)[:2000]

    # 2. Groundedness — full concatenated context vs answer
    full_context = "\n\n".join(context_texts)
    s, r = _call_provider(
        provider.groundedness_measure_with_cot_reasons, full_context, answer
    )
    scores["groundedness"]  = s
    reasons["groundedness"] = r

    # 3. Answer Relevance — question vs answer
    s, r = _call_provider(
        provider.relevance_with_cot_reasons, question, answer
    )
    scores["answer_relevance"]  = s
    reasons["answer_relevance"] = r

    return scores, reasons


# ─── PAGE ACCURACY (deterministic) ───────────────────────────────────────────

def score_page_accuracy(retrieved: list, ground_truth_pages: list):
    if not ground_truth_pages:
        return None
    retrieved_pages = set(c["page"] for c in retrieved if c.get("page"))
    hits = len(retrieved_pages & set(ground_truth_pages))
    return round(hits / len(ground_truth_pages), 3)


# ─── EXCEL REPORT ────────────────────────────────────────────────────────────

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E5D9E"
LIGHT_BLUE = "D9E1F2"
GREEN      = "E2EFDA"
AMBER      = "FFF2CC"
RED        = "FCE4D6"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"
GREEN_FONT = "375623"
AMBER_FONT = "7D6608"
RED_FONT   = "9C0006"


def _fill(c):
    return PatternFill("solid", start_color=c, fgColor=c)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(wrap=True, h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _score_color(score):
    if score is None:
        return GREY, "555555"
    if score >= 0.75:
        return GREEN, GREEN_FONT
    if score >= 0.40:
        return AMBER, AMBER_FONT
    return RED, RED_FONT


def build_report(results: list, output_path: str):
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    def write_title(ws, row, text, span, bg=DARK_BLUE, size=13):
        ws.merge_cells(span)
        c = ws[span.split(":")[0]]
        c.value     = text
        c.font      = _font(bold=True, color=WHITE, size=size)
        c.fill      = _fill(bg)
        c.alignment = _align(h="center", v="center", wrap=False)
        ws.row_dimensions[row].height = 26

    write_title(ws, 1, "RAG Evaluation Report — TruLens RAG Triad", "A1:H1")

    ws.merge_cells("A2:H2")
    ws["A2"].value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"RAG model: {GEN_MODEL} | Embeddings: {EMBED_MODEL} | "
        f"Judge: LiteLLM → ollama/{GEN_MODEL} | Questions: {len(results)}"
    )
    ws["A2"].font      = _font(color=WHITE, size=9)
    ws["A2"].fill      = _fill(MID_BLUE)
    ws["A2"].alignment = _align(h="center", wrap=False)
    ws.row_dimensions[2].height = 15

    ws.merge_cells("A4:H4")
    ws["A4"].value     = "TruLens RAG Triad — three feedback functions (all higher = better)"
    ws["A4"].font      = _font(bold=True, color=WHITE, size=10)
    ws["A4"].fill      = _fill(MID_BLUE)
    ws["A4"].alignment = _align(h="center", wrap=False)

    for ri, (metric, layer, equiv, desc) in enumerate([
        ("Context Relevance", "Retriever",
         "RAGAS: Context Precision | DeepEval: ContextualPrecision",
         "For each retrieved chunk, is it relevant to the input query? "
         "Scored per chunk via CoT reasoning then averaged across TOP_K chunks."),
        ("Groundedness", "Generator",
         "RAGAS: Faithfulness | DeepEval: Faithfulness + Hallucination",
         "Is the generated answer grounded in the retrieved context? "
         "The judge decomposes the answer into claims and checks each against "
         "the context with chain-of-thought reasoning. Higher = fewer hallucinations."),
        ("Answer Relevance", "End-to-end",
         "RAGAS: Answer Relevancy | DeepEval: AnswerRelevancy",
         "Does the final answer address the original question? "
         "Evaluates end-to-end utility of the RAG response."),
        ("Page Accuracy", "Retriever",
         "Custom (deterministic, same across RAGAS / DeepEval / TruLens)",
         "Did the retriever find a chunk from the correct page(s) of the "
         "Basel Framework? N/A for out-of-scope questions."),
    ], start=5):
        ws.cell(row=ri, column=1, value=metric).font = _font(bold=True, size=9)
        ws.cell(row=ri, column=2, value=layer)
        ws.cell(row=ri, column=3, value=equiv)
        ws.cell(row=ri, column=4, value=desc)
        ws.merge_cells(f"D{ri}:H{ri}")
        for ci in range(1, 5):
            c = ws.cell(row=ri, column=ci)
            c.fill      = _fill(LIGHT_BLUE if ri % 2 == 0 else WHITE)
            c.border    = _border()
            c.alignment = _align(wrap=True, v="center")
            if not c.font.bold:
                c.font = _font(size=9)
        ws.row_dimensions[ri].height = 42

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 60

    # Aggregate scores
    ws.row_dimensions[9].height = 10
    write_title(ws, 10, "Aggregate scores (mean across all questions)",
                "A10:H10", bg=MID_BLUE, size=10)

    for ci, (label, key) in enumerate([
        ("Context\nRelevance ↑", "context_relevance"),
        ("Groundedness ↑",       "groundedness"),
        ("Answer\nRelevance ↑",  "answer_relevance"),
        ("Page\nAccuracy ↑",     "page_accuracy"),
    ], start=1):
        vals = [r[key] for r in results if r.get(key) is not None]
        avg  = round(sum(vals) / len(vals), 3) if vals else None
        bg, fg = _score_color(avg)

        hc = ws.cell(row=11, column=ci, value=label)
        hc.font = _font(bold=True, size=8, color="555555")
        hc.fill = _fill(GREY); hc.alignment = _align(h="center", wrap=True)
        hc.border = _border()

        vc = ws.cell(row=12, column=ci, value=f"{avg:.2f}" if avg is not None else "N/A")
        vc.font = _font(bold=True, size=20, color=fg)
        vc.fill = _fill(bg); vc.alignment = _align(h="center", v="center", wrap=False)
        vc.border = _border()
        ws.row_dimensions[12].height = 40
        ws.column_dimensions[get_column_letter(ci)].width = 20

    # Framework comparison
    ws.row_dimensions[13].height = 10
    write_title(ws, 14, "TruLens vs RAGAS vs DeepEval — key difference",
                "A14:H14", bg=MID_BLUE, size=10)
    for ri, (fw, note) in enumerate([
        ("RAGAS",    "Free-form LLM prompt → parse number from text. No CoT. No tracing."),
        ("DeepEval", "Pydantic JSON output → typed score + reason. No tracing."),
        ("TruLens",  "LiteLLM CoT reasoning → score + step-by-step explanation. "
                     "Same judge LLM, most interpretable output of the three frameworks."),
    ], start=15):
        fc = ws.cell(row=ri, column=1, value=fw)
        dc = ws.cell(row=ri, column=2, value=note)
        ws.merge_cells(f"B{ri}:H{ri}")
        fc.font = _font(bold=True, size=9)
        dc.font = _font(size=9)
        for c in (fc, dc):
            c.fill = _fill(LIGHT_BLUE if ri % 2 == 0 else WHITE)
            c.border = _border(); c.alignment = _align(wrap=True, v="center")
        ws.row_dimensions[ri].height = 30

    # Score guide
    ws.row_dimensions[18].height = 10
    write_title(ws, 19, "Score interpretation (all metrics: higher = better)",
                "A19:H19", bg=MID_BLUE, size=10)
    for ri, (rng, label, bg, fg, advice) in enumerate([
        (">=0.75", "Good",  GREEN, GREEN_FONT, "Performing well."),
        ("0.40-0.74", "Fair", AMBER, AMBER_FONT, "Acceptable, investigate failures via Reasons sheet."),
        ("<0.40",  "Poor",  RED,   RED_FONT,   "Systematic issue — check CHUNK_SIZE, TOP_K, or embedding model."),
    ], start=20):
        rc = ws.cell(row=ri, column=1, value=rng)
        lc = ws.cell(row=ri, column=2, value=label)
        ac = ws.cell(row=ri, column=3, value=advice)
        ws.merge_cells(f"C{ri}:H{ri}")
        rc.font = _font(bold=True, size=10, color=fg)
        lc.font = _font(bold=True, size=10, color=fg)
        ac.font = _font(size=9)
        for c in (rc, lc, ac):
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _align(h="center" if c is not ac else "left", wrap=False)
        ws.row_dimensions[ri].height = 18

    # ── Sheet 2: Per-question scores ─────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Question Scores")
    headers    = ["Question", "Context Relevance ↑", "Groundedness ↑",
                  "Answer Relevance ↑", "Page Accuracy ↑",
                  "Pages Retrieved", "Pages Expected", "Time (s)"]
    col_widths = [38, 16, 16, 16, 14, 16, 16, 10]
    score_cols = {2, 3, 4, 5}

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center", v="center", wrap=True)
        c.border = _border()
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 32
    ws2.freeze_panes = "A2"

    for ri, result in enumerate(results, start=2):
        row_bg   = LIGHT_BLUE if ri % 2 == 0 else WHITE
        row_data = [
            result["question"],
            result.get("context_relevance"),
            result.get("groundedness"),
            result.get("answer_relevance"),
            result.get("page_accuracy"),
            ", ".join(f"p.{p}" for p in result.get("retrieved_pages", [])) or "—",
            ", ".join(f"p.{p}" for p in result.get("ground_truth_pages", [])) or "—",
            f"{result.get('elapsed_s', 0):.1f}",
        ]
        for ci, val in enumerate(row_data, start=1):
            c = ws2.cell(row=ri, column=ci, value=val if val is not None else "N/A")
            c.border = _border(); c.alignment = _align(wrap=True); c.font = _font(size=9)
            if ci in score_cols and isinstance(val, float):
                bg, fg = _score_color(val)
                c.fill = _fill(bg); c.font = _font(bold=True, color=fg, size=9)
                c.value = f"{val:.2f}"
            elif ci in score_cols:
                c.fill = _fill(GREY)
            else:
                c.fill = _fill(row_bg)
        ws2.row_dimensions[ri].height = 60

    # ── Sheet 3: Reasons (CoT explanations) ──────────────────────────────────
    ws3 = wb.create_sheet("Reasons")
    ws3.merge_cells("A1:D1")
    ws3["A1"].value     = "TruLens — Chain-of-Thought reasoning for every score"
    ws3["A1"].font      = _font(bold=True, color=WHITE, size=11)
    ws3["A1"].fill      = _fill(DARK_BLUE)
    ws3["A1"].alignment = _align(h="center", wrap=False)
    ws3.row_dimensions[1].height = 22

    for ci, (h, w) in enumerate(zip(
        ["Question", "Context Relevance Reasoning",
         "Groundedness Reasoning", "Answer Relevance Reasoning"],
        [38, 55, 55, 55]
    ), start=1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(MID_BLUE); c.border = _border()
        c.alignment = _align(h="center", wrap=True)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[2].height = 20
    ws3.freeze_panes = "A3"

    for ri, result in enumerate(results, start=3):
        row_bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        c = ws3.cell(row=ri, column=1, value=result["question"])
        c.fill = _fill(row_bg); c.border = _border()
        c.alignment = _align(wrap=True); c.font = _font(size=9)
        for ci, key in enumerate(
            ["context_relevance", "groundedness", "answer_relevance"], start=2
        ):
            c = ws3.cell(row=ri, column=ci,
                         value=result.get("reasons", {}).get(key, "N/A"))
            c.font = _font(size=9); c.fill = _fill(row_bg)
            c.border = _border(); c.alignment = _align(wrap=True)
        ws3.row_dimensions[ri].height = 100

    # ── Sheet 4: Answers ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Answers")
    ws4.merge_cells("A1:D1")
    ws4["A1"].value     = "Generated answers vs expected outputs — qualitative review"
    ws4["A1"].font      = _font(bold=True, color=WHITE, size=11)
    ws4["A1"].fill      = _fill(DARK_BLUE)
    ws4["A1"].alignment = _align(h="center", wrap=False)

    for ci, (h, w) in enumerate(zip(
        ["Question", "Generated Answer", "Expected Output", "Pages Retrieved"],
        [38, 60, 60, 20]
    ), start=1):
        c = ws4.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(MID_BLUE); c.border = _border()
        c.alignment = _align(h="center", wrap=True)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.freeze_panes = "A3"

    for ri, result in enumerate(results, start=3):
        row_bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        for ci, val in enumerate([
            result["question"],
            result.get("answer", ""),
            result.get("expected_output", ""),
            ", ".join(f"p.{p}" for p in result.get("retrieved_pages", [])) or "—",
        ], start=1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.font = _font(size=9); c.fill = _fill(row_bg)
            c.border = _border(); c.alignment = _align(wrap=True)
        ws4.row_dimensions[ri].height = 90

    # ── Sheet 5: Config ──────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Config")
    for ri, (k, v) in enumerate([
        ("Parameter",        "Value"),
        ("OLLAMA_URL",        OLLAMA_URL),
        ("GEN_MODEL",         GEN_MODEL),
        ("EMBED_MODEL",       EMBED_MODEL),
        ("Judge",             f"LiteLLM → ollama/{GEN_MODEL}"),
        ("CHUNK_SIZE",        CHUNK_SIZE),
        ("CHUNK_OVERLAP",     CHUNK_OVERLAP),
        ("TOP_K",             TOP_K),
        ("SCORE_THRESHOLD",   SCORE_THRESHOLD),
        ("RRF_K",             RRF_K),
        ("Questions",         len(results)),
        ("Ollama endpoint",   "/api/embed (Ollama >= 0.2)"),
        ("TruLens version",   "2.8.1"),
        ("Framework",         "TruLens (Snowflake)"),
        ("Run timestamp",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ], start=1):
        ck = ws5.cell(row=ri, column=1, value=k)
        cv = ws5.cell(row=ri, column=2, value=v)
        if ri == 1:
            for c in (ck, cv):
                c.font = _font(bold=True, color=WHITE); c.fill = _fill(DARK_BLUE)
        else:
            bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
            for c in (ck, cv):
                c.fill = _fill(bg); c.font = _font(size=10)
        for c in (ck, cv):
            c.border = _border(); c.alignment = _align(wrap=False)
    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 55

    wb.save(output_path)
    print(f"\n  Report saved → {output_path}")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    print("\n=== RAG Evaluation — TruLens RAG Triad (5-step scientific method) ===\n")

    print("STEP 1 — OBSERVATION")
    print("  LLMs hallucinate on unseen regulatory documents.")
    print("  Our hybrid RAG over the Basel Framework should reduce this.")
    print("  TruLens verifies with the RAG Triad + CoT reasoning per score.\n")

    print("STEP 2 — HYPOTHESIS")
    print(f"  RAG   : {GEN_MODEL} + {EMBED_MODEL}, TOP_K={TOP_K}, RRF_K={RRF_K}")
    print(f"  Judge : LiteLLM → ollama/{GEN_MODEL} — fully local, no OpenAI key")
    print(f"  Triad : Context Relevance ↑  Groundedness ↑  Answer Relevance ↑")
    print(f"  CoT   : Judge explains reasoning before assigning each score\n")

    print("STEP 3 — EXPERIMENT")
    print(f"  [1/3] Loading documents from '{DOCS_DIR}'...")
    docs = load_documents(DOCS_DIR)
    if not docs:
        print("  No documents found. Add PDFs to ./docs/ and rerun.")
        return
    chunks = chunk_documents(docs)
    print(f"  {len(chunks)} chunks ready.\n")

    print("  [2/3] Setting up indexes...")
    collection = get_collection()
    index_chunks(collection, chunks)
    bm25_index = BM25Index(chunks)

    print(f"\n  [3/3] Evaluating {len(GROUND_TRUTH)} questions...\n")
    header = (f"  {'Question':<40} {'CtxRel':>7} {'Ground':>7} "
              f"{'AnsRel':>7} {'Page':>6} {'Time':>6}")
    print(header)
    print("  " + "-" * (len(header) - 2))

    results = []
    for entry in GROUND_TRUTH:
        question        = entry["question"]
        expected_output = entry["expected_output"]
        gt_pages        = entry["ground_truth_pages"]

        t0        = time.time()
        retrieved = hybrid_retrieve(question, collection, bm25_index)
        answer    = generate_answer(question, retrieved)
        scores, reasons = run_trulens_metrics(question, answer, retrieved)
        pa        = score_page_accuracy(retrieved, gt_pages)
        elapsed   = round(time.time() - t0, 1)

        def _f(v):
            return f"{v:7.2f}" if v is not None else "    N/A"

        print(
            f"  {question[:40]:<40} "
            f"{_f(scores.get('context_relevance')):>7} "
            f"{_f(scores.get('groundedness')):>7} "
            f"{_f(scores.get('answer_relevance')):>7} "
            f"{'N/A' if pa is None else f'{pa:5.2f}':>6} "
            f"{elapsed}s"
        )

        results.append({
            "question":           question,
            "answer":             answer,
            "expected_output":    expected_output,
            "ground_truth_pages": gt_pages,
            "retrieved_pages":    sorted(set(
                c["page"] for c in retrieved if c.get("page")
            )),
            **scores,
            "page_accuracy": pa,
            "reasons":       reasons,
            "elapsed_s":     elapsed,
        })

    print("\nSTEP 4 — ANALYSIS\n")
    def mean(key):
        vals = [r[key] for r in results if r.get(key) is not None]
        return round(sum(vals) / len(vals), 3) if vals else None

    for label, key in [
        ("Context Relevance  (↑)", "context_relevance"),
        ("Groundedness       (↑)", "groundedness"),
        ("Answer Relevance   (↑)", "answer_relevance"),
        ("Page Accuracy      (↑)", "page_accuracy"),
    ]:
        val = mean(key)
        if val is None:
            print(f"  {label}: N/A")
            continue
        tag = "GOOD" if val >= 0.75 else "FAIR" if val >= 0.40 else "POOR"
        print(f"  {label}: {val:.3f}  [{tag}]")

    print("\nSTEP 5 — CONCLUSION")
    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"trulens_report_{ts}.xlsx"
    build_report(results, output_path)
    print("  Open 'Reasons' sheet for CoT explanations per score.")
    print("  Open 'Answers' sheet to compare generated vs expected outputs.\n")


if __name__ == "__main__":
    main()
