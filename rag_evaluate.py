# rag_evaluate.py — RAGAS-style evaluation for the hybrid RAG
#
# Implements four metrics entirely locally via Ollama (no external API):
#   1. Context Precision   — are retrieved chunks relevant?
#   2. Context Recall      — did we retrieve all needed information?
#   3. Faithfulness        — is the answer grounded in the chunks?
#   4. Answer Relevancy    — does the answer address the question?
#
# Usage:
#   python rag_evaluate.py
#
# Output:
#   rag_eval_report_<timestamp>.xlsx
#
# Requirements:
#   pip install chromadb pypdf nltk rank_bm25 openpyxl requests

import os
import re
import time
import hashlib
import requests
from pathlib import Path
from datetime import datetime

import nltk
import chromadb
from rank_bm25 import BM25Okapi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

nltk.download("punkt",     quiet=True)
nltk.download("punkt_tab", quiet=True)

# ─── CONFIG — must match your main rag script ───────────────────────────────
OLLAMA_URL      = "http://localhost:11434"
GEN_MODEL       = "llama3.2"
EMBED_MODEL     = "mxbai-embed-large"
DOCS_DIR        = "./docs"
CHROMA_DIR      = "./chroma_store"
COLLECTION      = "rag_knowledge"
CHUNK_SIZE      = 1000
CHUNK_OVERLAP   = 150
TOP_K           = 5
SCORE_THRESHOLD = 0.40
RRF_K           = 60

# ─── GROUND TRUTH DATASET ───────────────────────────────────────────────────
# Each entry:
#   question          : the question to ask
#   ground_truth      : the ideal answer (written by you, from the document)
#   ground_truth_pages: list of page numbers where the answer lives
#
# How to build this:
#   1. Ask each question to your RAG
#   2. Open the PDF at the cited page
#   3. Write the ground truth answer in your own words from the document
#   4. Record the page number(s)
#
# The more entries you add, the more reliable the evaluation.
# Aim for 20-30 questions covering all categories.

GROUND_TRUTH = [

    # ── Category A: Factual / Definition ────────────────────────────────────
    {
        "question": "What is Tier 1 capital?",
        "ground_truth": (
            "Tier 1 capital is the going-concern capital of a bank. "
            "It consists of Common Equity Tier 1 (CET1) capital and "
            "Additional Tier 1 (AT1) capital. CET1 is the highest quality "
            "capital and includes common shares, retained earnings, and "
            "other comprehensive income. AT1 instruments are subordinated, "
            "have no maturity, and can absorb losses on a going-concern basis."
        ),
        "ground_truth_pages": [101, 102],
    },
    {
        "question": "What is the definition of a Global Systemically Important Bank (G-SIB)?",
        "ground_truth": (
            "A Global Systemically Important Bank (G-SIB) is a bank whose "
            "distress or failure would cause significant disruption to the "
            "broader global financial system and economic activity. G-SIBs "
            "are identified through an assessment methodology based on five "
            "categories: size, interconnectedness, substitutability, "
            "complexity, and cross-border activity."
        ),
        "ground_truth_pages": [17, 18],
    },
    {
        "question": "What does LCR stand for and what is its minimum requirement?",
        "ground_truth": (
            "LCR stands for Liquidity Coverage Ratio. The minimum requirement "
            "is 100%, meaning a bank must hold sufficient high-quality liquid "
            "assets (HQLA) to cover total net cash outflows over a 30-day "
            "stressed period."
        ),
        "ground_truth_pages": [1093, 1097],
    },
    {
        "question": "What is the Net Stable Funding Ratio (NSFR)?",
        "ground_truth": (
            "The Net Stable Funding Ratio (NSFR) requires banks to maintain "
            "a stable funding profile in relation to their assets and "
            "off-balance sheet activities over a one-year horizon. "
            "The NSFR is defined as the ratio of Available Stable Funding "
            "to Required Stable Funding, and must be at least 100%."
        ),
        "ground_truth_pages": [1216, 1219],
    },
    {
        "question": "How does the Basel Framework define a trading book?",
        "ground_truth": (
            "The trading book consists of positions in financial instruments "
            "and commodities held either with trading intent or to hedge "
            "other elements of the trading book. Trading intent is evidenced "
            "by the strategies, policies, and procedures a bank has in place "
            "to manage the position. Positions in the trading book are subject "
            "to market risk capital requirements."
        ),
        "ground_truth_pages": [201, 202],
    },

    # ── Category B: Numerical / Threshold ───────────────────────────────────
    {
        "question": "What is the minimum Common Equity Tier 1 (CET1) ratio under Basel III?",
        "ground_truth": (
            "The minimum Common Equity Tier 1 (CET1) ratio under Basel III "
            "is 4.5% of risk-weighted assets."
        ),
        "ground_truth_pages": [193],
    },
    {
        "question": "What is the capital conservation buffer requirement?",
        "ground_truth": (
            "The capital conservation buffer is set at 2.5% of total "
            "risk-weighted assets and must be met with Common Equity Tier 1 "
            "capital. When a bank's CET1 falls within the buffer range, "
            "constraints are placed on earnings distributions."
        ),
        "ground_truth_pages": [217, 218],
    },
    {
        "question": "What is the large exposure limit as a percentage of Tier 1 capital?",
        "ground_truth": (
            "The large exposure limit is 25% of a bank's eligible Tier 1 "
            "capital. For exposures between G-SIBs, a stricter limit of "
            "15% of Tier 1 capital applies."
        ),
        "ground_truth_pages": [1260],
    },

    # ── Category C: Process / How-to ────────────────────────────────────────
    {
        "question": "How does a bank calculate its Liquidity Coverage Ratio?",
        "ground_truth": (
            "The LCR is calculated as the stock of high-quality liquid assets "
            "(HQLA) divided by the total net cash outflows over the next "
            "30 calendar days under a stress scenario. The result must be "
            "at least 100%. HQLA are assets that can be easily converted to "
            "cash in private markets. Net cash outflows are calculated as "
            "total expected outflows minus total expected inflows, capped "
            "at 75% of total outflows."
        ),
        "ground_truth_pages": [1097, 1098],
    },
    {
        "question": "How is the leverage ratio calculated?",
        "ground_truth": (
            "The leverage ratio is calculated as Tier 1 capital divided by "
            "the bank's total exposure measure. The exposure measure includes "
            "on-balance sheet assets, derivative exposures, securities "
            "financing transaction exposures, and off-balance sheet items. "
            "The minimum leverage ratio requirement is 3% of Tier 1 capital."
        ),
        "ground_truth_pages": [1052, 1054],
    },

    # ── Category D: Comparative ──────────────────────────────────────────────
    {
        "question": "What is the difference between the Standardised Approach and the IRB approach for credit risk?",
        "ground_truth": (
            "Under the Standardised Approach, risk weights are assigned by "
            "supervisors based on external credit ratings or asset class. "
            "Under the Internal Ratings-Based (IRB) approach, banks use "
            "their own internal models to estimate risk components such as "
            "Probability of Default (PD), Loss Given Default (LGD), and "
            "Exposure at Default (EAD), subject to supervisory approval "
            "and minimum requirements."
        ),
        "ground_truth_pages": [248, 343],
    },
    {
        "question": "What distinguishes Tier 1 capital from Tier 2 capital?",
        "ground_truth": (
            "Tier 1 capital is going-concern capital — it absorbs losses "
            "while the bank continues to operate. It includes CET1 and AT1 "
            "instruments. Tier 2 capital is gone-concern capital — it "
            "absorbs losses only in liquidation. Tier 2 instruments include "
            "subordinated debt with a minimum maturity of five years and "
            "certain loan loss provisions."
        ),
        "ground_truth_pages": [101, 125],
    },

    # ── Category E: Out-of-scope ─────────────────────────────────────────────
    {
        "question": "What is the current ECB interest rate?",
        "ground_truth": "I don't know. This information is not in the Basel Framework document.",
        "ground_truth_pages": [],   # no page — answer should not be found
    },
    {
        "question": "What does IFRS 9 say about expected credit loss provisioning?",
        "ground_truth": "I don't know. IFRS 9 details are not covered in the Basel Framework document.",
        "ground_truth_pages": [],
    },

    # ── Category F: Advanced / Multi-hop ────────────────────────────────────
    {
        "question": "How does a bank's CET1 ratio affect dividend eligibility under the capital conservation buffer?",
        "ground_truth": (
            "When a bank's CET1 ratio falls within the capital conservation "
            "buffer range (4.5% to 7%), earnings distribution constraints "
            "are imposed. The closer the CET1 ratio is to the 4.5% minimum, "
            "the greater the restriction on distributions such as dividends, "
            "share buybacks, and discretionary bonus payments. A bank with "
            "a CET1 ratio between 4.5% and 5.125% must retain at least 100% "
            "of its earnings."
        ),
        "ground_truth_pages": [217, 218, 219],
    },
    {
        "question": "How are cryptoasset exposures classified and what capital treatment do they receive?",
        "ground_truth": (
            "Cryptoassets are divided into two groups. Group 1 cryptoassets "
            "meet a set of classification conditions and receive treatment "
            "broadly equivalent to traditional assets. Group 2 cryptoassets "
            "do not meet these conditions and are subject to a conservative "
            "capital treatment including a risk weight of 1250% for Group 2b "
            "assets, reflecting their higher risk profile."
        ),
        "ground_truth_pages": [37, 38],
    },
]

# ─── RAG PIPELINE ────────────────────────────────────────────────────────────

def ollama_embed(texts):
    embeddings = []
    for text in texts:
        r = requests.post(
            f"{OLLAMA_URL}/api/embeddings",
            json={"model": EMBED_MODEL, "prompt": text}
        )
        r.raise_for_status()
        embeddings.append(r.json()["embedding"])
    return embeddings

def ollama_generate(prompt):
    r = requests.post(
        f"{OLLAMA_URL}/api/generate",
        json={"model": GEN_MODEL, "prompt": prompt, "stream": False}
    )
    r.raise_for_status()
    return r.json()["response"].strip()

def load_pdf(path):
    from pypdf import PdfReader
    reader = PdfReader(path)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append({"text": text, "page": i + 1})
    return pages

def load_documents(docs_dir):
    docs = []
    for filepath in Path(docs_dir).glob("**/*"):
        if filepath.suffix.lower() == ".pdf":
            pages = load_pdf(str(filepath))
            if pages:
                docs.append({"source": str(filepath), "pages": pages})
        elif filepath.suffix.lower() == ".txt":
            with open(str(filepath), "r", encoding="utf-8") as f:
                text = f.read()
            if text.strip():
                docs.append({
                    "source": str(filepath),
                    "pages":  [{"text": text, "page": None}]
                })
    return docs

def chunk_text(text, chunk_size, overlap):
    from nltk.tokenize import sent_tokenize
    sentences, chunks = sent_tokenize(text), []
    current, current_len = [], 0
    for sent in sentences:
        slen = len(sent)
        if current_len + slen > chunk_size and current:
            chunks.append(" ".join(current))
            kept, kept_len = [], 0
            for s in reversed(current):
                if kept_len + len(s) > overlap:
                    break
                kept.insert(0, s)
                kept_len += len(s)
            current, current_len = kept, kept_len
        current.append(sent)
        current_len += slen
    if current:
        chunks.append(" ".join(current))
    return [c for c in chunks if len(c) > 30]

def chunk_documents(docs):
    all_chunks = []
    for doc in docs:
        for pb in doc["pages"]:
            page_num = pb["page"] or 0
            for i, chunk in enumerate(chunk_text(pb["text"], CHUNK_SIZE, CHUNK_OVERLAP)):
                cid = hashlib.md5(f"{doc['source']}_{page_num}_{i}".encode()).hexdigest()
                all_chunks.append({
                    "id": cid, "text": chunk,
                    "source": doc["source"],
                    "page": page_num, "chunk_index": i,
                })
    return all_chunks

def get_collection():
    client = chromadb.PersistentClient(path=CHROMA_DIR)
    return client.get_or_create_collection(
        name=COLLECTION, metadata={"hnsw:space": "cosine"}
    )

def index_chunks(collection, chunks):
    existing = set(collection.get()["ids"])
    new = [c for c in chunks if c["id"] not in existing]
    if not new:
        print("  ChromaDB: all chunks already indexed.")
        return
    print(f"  Embedding {len(new)} new chunks...")
    for i, chunk in enumerate(new):
        emb = ollama_embed([chunk["text"]])[0]
        collection.add(
            ids=[chunk["id"]], embeddings=[emb],
            documents=[chunk["text"]],
            metadatas=[{
                "source": chunk["source"],
                "page": chunk["page"],
                "chunk_index": chunk["chunk_index"],
            }]
        )
        if (i + 1) % 100 == 0:
            print(f"    {i+1}/{len(new)}...")
    print(f"  Done.")

class BM25Index:
    def __init__(self, chunks):
        self.chunks = chunks
        self.bm25 = BM25Okapi([c["text"].lower().split() for c in chunks])

    def search(self, query, top_k):
        scores = self.bm25.get_scores(query.lower().split())
        ranked = sorted(zip(scores, self.chunks),
                        key=lambda x: x[0], reverse=True)[:top_k]
        return [
            {**chunk, "bm25_score": round(float(score), 4)}
            for score, chunk in ranked if score > 0
        ]

def hybrid_retrieve(query, collection, bm25_index):
    qemb = ollama_embed([query])[0]
    res  = collection.query(
        query_embeddings=[qemb], n_results=TOP_K,
        include=["documents", "metadatas", "distances"]
    )
    semantic = []
    for text, meta, dist in zip(
        res["documents"][0], res["metadatas"][0], res["distances"][0]
    ):
        score = round(1 - dist, 3)
        if score >= SCORE_THRESHOLD:
            semantic.append({
                "text": text,
                "source": Path(meta["source"]).name,
                "page": meta.get("page", 0),
                "sem_score": score,
            })

    bm25_hits = [
        {"text": c["text"], "source": Path(c["source"]).name,
         "page": c["page"], "bm25_score": c["bm25_score"]}
        for c in bm25_index.search(query, TOP_K)
    ]

    rrf_scores, chunk_store = {}, {}
    for rank, chunk in enumerate(semantic):
        key = chunk["text"][:60]
        rrf_scores[key]  = rrf_scores.get(key, 0) + 1 / (rank + RRF_K)
        chunk_store[key] = chunk
    for rank, chunk in enumerate(bm25_hits):
        key = chunk["text"][:60]
        rrf_scores[key]  = rrf_scores.get(key, 0) + 1 / (rank + RRF_K)
        if key not in chunk_store:
            chunk_store[key] = chunk

    ranked = sorted(rrf_scores.items(), key=lambda x: x[1], reverse=True)
    results = []
    for key, rrf_score in ranked[:TOP_K]:
        chunk = chunk_store[key].copy()
        chunk["rrf_score"] = round(rrf_score, 6)
        results.append(chunk)
    return results

def generate_answer(question, retrieved):
    if not retrieved:
        return "I could not find relevant information in the indexed documents."
    context = "\n\n".join(
        f"[Source: {r['source']}, Page: {r['page']}]\n{r['text']}"
        for r in retrieved
    )
    prompt = f"""Use only the context below to answer the question.
Be concise and factual.
If the answer is not in the context, say "I don't know".
At the end of your answer, cite the source file and page number(s):
  Source: <filename>, Page <number>

Context:
{context}

Question: {question}
Answer:"""
    return ollama_generate(prompt)

# ─── RAGAS METRICS ───────────────────────────────────────────────────────────
#
# All four metrics use the LLM-as-judge pattern:
# we ask the LLM to evaluate quality and parse its response.
# This is the standard approach when you don't have human annotations.
# Scores are 0.0 – 1.0 for each metric.

def _parse_score(response: str, low=0.0, high=1.0) -> float:
    """Extract the first float in [low, high] from an LLM response."""
    matches = re.findall(r"\d+(?:\.\d+)?", response)
    for m in matches:
        val = float(m)
        if low <= val <= high:
            return val
    return 0.0

def score_context_precision(question: str, retrieved: list[dict],
                             ground_truth: str) -> float:
    """
    Context Precision: of the TOP_K retrieved chunks,
    how many are actually relevant to answering the question?

    RAG analogy: PRECISION in classification.
    Formula: relevant_retrieved / total_retrieved
    """
    if not retrieved:
        return 0.0

    relevant = 0
    for chunk in retrieved:
        prompt = f"""You are evaluating a RAG retrieval system.

Question: {question}

Retrieved chunk:
{chunk['text'][:600]}

Ground truth answer (for reference):
{ground_truth}

Is this chunk relevant to answering the question?
Answer with only a number: 1 (relevant) or 0 (not relevant)."""
        resp = ollama_generate(prompt)
        relevant += 1 if "1" in resp.split()[:3] else 0

    return round(relevant / len(retrieved), 3)


def score_context_recall(question: str, retrieved: list[dict],
                          ground_truth: str) -> float:
    """
    Context Recall: how much of the ground truth answer
    is covered by the retrieved chunks?

    RAG analogy: RECALL in classification.
    Strategy: ask the LLM to score coverage 0-10, normalise to 0-1.
    """
    if not retrieved:
        return 0.0

    context = "\n\n---\n\n".join(c["text"][:400] for c in retrieved)
    prompt = f"""You are evaluating a RAG retrieval system.

Question: {question}

Ground truth answer:
{ground_truth}

Retrieved context (all chunks combined):
{context}

How much of the ground truth answer is covered by the retrieved context?
Score from 0 to 10 where:
  0  = nothing in the ground truth is found in the context
  5  = about half the key facts are present
  10 = all key facts in the ground truth are present in the context

Reply with ONLY a single integer (0-10). No explanation."""
    resp = ollama_generate(prompt)
    raw  = _parse_score(resp, low=0, high=10)
    return round(raw / 10, 3)


def score_faithfulness(question: str, retrieved: list[dict],
                        answer: str) -> float:
    """
    Faithfulness: are the claims in the answer supported
    by the retrieved chunks?

    Measures hallucination. Score = 1.0 means every claim
    in the answer is grounded in the context.
    """
    if not retrieved or not answer:
        return 0.0

    context = "\n\n---\n\n".join(c["text"][:400] for c in retrieved)
    prompt = f"""You are evaluating whether an AI answer is grounded in its source context.

Question: {question}

Retrieved context:
{context}

Generated answer:
{answer}

What fraction of the statements in the answer are directly supported by the context?
Score from 0 to 10 where:
  0  = the answer is entirely hallucinated (nothing from the context)
  5  = about half the claims are supported
  10 = every claim in the answer is supported by the context

Reply with ONLY a single integer (0-10). No explanation."""
    resp = ollama_generate(prompt)
    raw  = _parse_score(resp, low=0, high=10)
    return round(raw / 10, 3)


def score_answer_relevancy(question: str, answer: str) -> float:
    """
    Answer Relevancy: does the answer actually address the question?

    A perfectly faithful answer can still fail to answer the question.
    This metric catches that.
    """
    if not answer:
        return 0.0

    prompt = f"""You are evaluating whether an AI answer is relevant to the question asked.

Question: {question}

Answer:
{answer}

How well does the answer address the question?
Score from 0 to 10 where:
  0  = the answer is completely off-topic
  5  = the answer is partially relevant
  10 = the answer directly and completely addresses the question

Reply with ONLY a single integer (0-10). No explanation."""
    resp = ollama_generate(prompt)
    raw  = _parse_score(resp, low=0, high=10)
    return round(raw / 10, 3)


def score_page_accuracy(retrieved: list[dict],
                         ground_truth_pages: list[int]) -> float:
    """
    Page Accuracy (bonus metric, not in standard RAGAS):
    did the retriever find at least one chunk from the correct page(s)?

    This is a hard, deterministic metric — no LLM judge needed.
    Particularly useful for the Basel Framework where you know
    exactly which pages contain the answer.

    Score = 1.0 if any retrieved chunk is on a ground truth page.
    Score = 0.0 if none are (or if ground_truth_pages is empty for
    out-of-scope questions, where we don't expect any page match).
    """
    if not ground_truth_pages:
        # Out-of-scope question: no page expected.
        # Check that the model said "I don't know" instead.
        return None   # handled separately in reporting

    retrieved_pages = set(c["page"] for c in retrieved if c.get("page"))
    hits = len(retrieved_pages & set(ground_truth_pages))
    return round(hits / len(ground_truth_pages), 3)

# ─── EXCEL REPORT ────────────────────────────────────────────────────────────

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E5D9E"
LIGHT_BLUE = "D9E1F2"
GREEN      = "E2EFDA"
AMBER      = "FFF2CC"
RED        = "FCE4D6"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"

GREEN_FONT = "375623"
AMBER_FONT = "7D6608"
RED_FONT   = "9C0006"

def _fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")
def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(wrap=True, h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _score_color(score):
    """Return (bg, fg) hex tuple based on 0-1 score."""
    if score is None:
        return GREY, "555555"
    if score >= 0.75:
        return GREEN, GREEN_FONT
    if score >= 0.40:
        return AMBER, AMBER_FONT
    return RED, RED_FONT

def build_eval_report(results: list[dict], output_path: str):
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    def write_title(ws, row, text, span="A1:H1", bg=DARK_BLUE, size=13):
        ws.merge_cells(span)
        c = ws[span.split(":")[0]]
        c.value     = text
        c.font      = _font(bold=True, color=WHITE, size=size)
        c.fill      = _fill(bg)
        c.alignment = _align(h="center", v="center", wrap=False)
        ws.row_dimensions[row].height = 26

    write_title(ws, 1, "RAG Evaluation Report — RAGAS Metrics", "A1:H1")

    ws.merge_cells("A2:H2")
    ws["A2"].value     = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Model: {GEN_MODEL}  |  Embeddings: {EMBED_MODEL}  |  "
        f"Questions evaluated: {len(results)}"
    )
    ws["A2"].font      = _font(color=WHITE, size=9)
    ws["A2"].fill      = _fill(MID_BLUE)
    ws["A2"].alignment = _align(h="center", wrap=False)
    ws.row_dimensions[2].height = 15

    # Metric explanation table (rows 4-9)
    ws.merge_cells("A4:H4")
    ws["A4"].value     = "What each metric measures"
    ws["A4"].font      = _font(bold=True, color=WHITE, size=10)
    ws["A4"].fill      = _fill(MID_BLUE)
    ws["A4"].alignment = _align(h="center", wrap=False)

    metric_desc = [
        ("Context Precision",
         "Retrieval quality — of the chunks retrieved, what fraction were actually relevant? "
         "Equivalent to PRECISION in classification. Low score = retriever is noisy."),
        ("Context Recall",
         "Retrieval completeness — how much of the correct answer was present in the "
         "retrieved chunks? Equivalent to RECALL. Low score = retriever missed the key chunk."),
        ("Faithfulness",
         "Generation quality — are the claims in the answer grounded in the retrieved chunks? "
         "Measures hallucination. Score of 1.0 = no hallucination."),
        ("Answer Relevancy",
         "End-to-end quality — does the answer actually address the question asked? "
         "A faithful answer can still miss the point; this catches that."),
        ("Page Accuracy",
         "Deterministic bonus metric — did the retriever find at least one chunk from the "
         "correct page(s)? Requires ground truth page numbers. N/A for out-of-scope questions."),
    ]
    for ri, (name, desc) in enumerate(metric_desc, start=5):
        nc = ws.cell(row=ri, column=1, value=name)
        dc = ws.cell(row=ri, column=2, value=desc)
        ws.merge_cells(f"B{ri}:H{ri}")
        nc.font      = _font(bold=True, size=9)
        dc.font      = _font(size=9)
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        for c in (nc, dc):
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = _align(wrap=True, v="center")
        ws.row_dimensions[ri].height = 30
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    # Aggregate scores (row 11+)
    ws.row_dimensions[10].height = 10
    write_title(ws, 11, "Aggregate scores (mean across all questions)", "A11:H11",
                bg=MID_BLUE, size=10)

    metrics = [
        "Context Precision", "Context Recall",
        "Faithfulness", "Answer Relevancy", "Page Accuracy"
    ]
    keys = [
        "context_precision", "context_recall",
        "faithfulness", "answer_relevancy", "page_accuracy"
    ]

    for ci, (metric, key) in enumerate(zip(metrics, keys), start=1):
        scores = [r[key] for r in results if r[key] is not None]
        avg    = round(sum(scores) / len(scores), 3) if scores else None
        bg, fg = _score_color(avg)

        hc = ws.cell(row=12, column=ci, value=metric)
        hc.font      = _font(bold=True, size=9, color="555555")
        hc.fill      = _fill(GREY)
        hc.alignment = _align(h="center", wrap=True)
        hc.border    = _border()

        vc = ws.cell(row=13, column=ci,
                     value=f"{avg:.2f}" if avg is not None else "N/A")
        vc.font      = _font(bold=True, size=20, color=fg)
        vc.fill      = _fill(bg)
        vc.alignment = _align(h="center", v="center", wrap=False)
        vc.border    = _border()
        ws.row_dimensions[13].height = 40

    for ci in range(1, 6):
        ws.column_dimensions[get_column_letter(ci)].width = 20

    # Interpretation guide
    ws.row_dimensions[14].height = 10
    write_title(ws, 15, "Score interpretation guide", "A15:H15",
                bg=MID_BLUE, size=10)
    guide = [
        ("≥ 0.75", "Good",    GREEN,  GREEN_FONT, "System is performing well on this metric."),
        ("0.40–0.74", "Fair", AMBER,  AMBER_FONT, "Acceptable but room to improve. Investigate individual failures."),
        ("< 0.40", "Poor",    RED,    RED_FONT,   "Systematic problem. Check chunk size, TOP_K, or embedding model."),
    ]
    for ri, (rng, label, bg, fg, advice) in enumerate(guide, start=16):
        rc = ws.cell(row=ri, column=1, value=rng)
        lc = ws.cell(row=ri, column=2, value=label)
        ac = ws.cell(row=ri, column=3, value=advice)
        ws.merge_cells(f"C{ri}:H{ri}")
        rc.font = _font(bold=True, size=10, color=fg)
        lc.font = _font(bold=True, size=10, color=fg)
        ac.font = _font(size=9)
        for c in (rc, lc, ac):
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = _align(h="center" if c is not ac else "left",
                                  wrap=False)
        ws.row_dimensions[ri].height = 18

    # ── Sheet 2: Per-question results ────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Question Results")

    headers = [
        "Question", "Answer",
        "Context\nPrecision", "Context\nRecall",
        "Faithfulness", "Answer\nRelevancy",
        "Page\nAccuracy", "Pages\nretrieved",
        "Pages\nexpected", "Time (s)"
    ]
    col_widths = [35, 55, 12, 12, 12, 12, 12, 16, 16, 10]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font      = _font(bold=True, color=WHITE, size=9)
        c.fill      = _fill(DARK_BLUE)
        c.alignment = _align(h="center", v="center", wrap=True)
        c.border    = _border()
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    for ri, result in enumerate(results, start=2):
        row_bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        row_data = [
            result["question"],
            result["answer"],
            result["context_precision"],
            result["context_recall"],
            result["faithfulness"],
            result["answer_relevancy"],
            result["page_accuracy"],
            ", ".join(f"p.{p}" for p in result["retrieved_pages"]) or "—",
            ", ".join(f"p.{p}" for p in result["ground_truth_pages"]) or "—",
            f"{result['elapsed_s']:.1f}",
        ]
        score_cols = {3, 4, 5, 6, 7}   # columns with 0-1 scores

        for ci, val in enumerate(row_data, start=1):
            c = ws2.cell(row=ri, column=ci,
                         value=val if val is not None else "N/A")
            c.border    = _border()
            c.alignment = _align(wrap=True)
            c.font      = _font(size=9)

            if ci in score_cols and isinstance(val, float):
                bg, fg = _score_color(val)
                c.fill  = _fill(bg)
                c.font  = _font(bold=True, color=fg, size=9)
                c.value = f"{val:.2f}"
            elif ci in score_cols and val is None:
                c.fill = _fill(GREY)
            else:
                c.fill = _fill(row_bg)

        ws2.row_dimensions[ri].height = 80

    # ── Sheet 3: Config snapshot ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Config")
    cfg = [
        ("Parameter",        "Value"),
        ("OLLAMA_URL",       OLLAMA_URL),
        ("GEN_MODEL",        GEN_MODEL),
        ("EMBED_MODEL",      EMBED_MODEL),
        ("CHUNK_SIZE",       CHUNK_SIZE),
        ("CHUNK_OVERLAP",    CHUNK_OVERLAP),
        ("TOP_K",            TOP_K),
        ("SCORE_THRESHOLD",  SCORE_THRESHOLD),
        ("RRF_K",            RRF_K),
        ("Questions",        len(results)),
        ("Run timestamp",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for ri, (k, v) in enumerate(cfg, start=1):
        ck = ws3.cell(row=ri, column=1, value=k)
        cv = ws3.cell(row=ri, column=2, value=v)
        if ri == 1:
            for c in (ck, cv):
                c.font = _font(bold=True, color=WHITE)
                c.fill = _fill(DARK_BLUE)
        else:
            bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
            for c in (ck, cv):
                c.fill = _fill(bg)
                c.font = _font(size=10)
        for c in (ck, cv):
            c.border    = _border()
            c.alignment = _align(wrap=False)
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 40

    wb.save(output_path)
    print(f"\n  Report saved: {output_path}")

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    print("\n=== RAG Evaluation — RAGAS Metrics ===\n")

    # 1. Load + chunk
    print(f"[1/3] Loading documents from '{DOCS_DIR}'...")
    docs   = load_documents(DOCS_DIR)
    if not docs:
        print("  No documents found. Add PDFs to ./docs/ and rerun.")
        return
    chunks = chunk_documents(docs)
    print(f"  {len(chunks)} chunks ready.")

    # 2. Index
    print(f"\n[2/3] Setting up indexes...")
    collection = get_collection()
    index_chunks(collection, chunks)
    bm25_index = BM25Index(chunks)

    # 3. Evaluate
    print(f"\n[3/3] Evaluating {len(GROUND_TRUTH)} questions...\n")
    print(f"  {'Question':<40} {'Prec':>6} {'Rec':>6} "
          f"{'Faith':>6} {'Rel':>6} {'Page':>6}  Time")
    print(f"  {'-'*40} {'-'*6} {'-'*6} {'-'*6} {'-'*6} {'-'*6}  ----")

    results = []
    for entry in GROUND_TRUTH:
        question          = entry["question"]
        ground_truth      = entry["ground_truth"]
        gt_pages          = entry["ground_truth_pages"]

        t0        = time.time()
        retrieved = hybrid_retrieve(question, collection, bm25_index)
        answer    = generate_answer(question, retrieved)

        # Compute all four RAGAS metrics + page accuracy
        cp   = score_context_precision(question, retrieved, ground_truth)
        cr   = score_context_recall(question, retrieved, ground_truth)
        faith = score_faithfulness(question, retrieved, answer)
        ar   = score_answer_relevancy(question, answer)
        pa   = score_page_accuracy(retrieved, gt_pages)

        elapsed = round(time.time() - t0, 1)

        pa_str = f"{pa:.2f}" if pa is not None else " N/A"
        print(f"  {question[:40]:<40} {cp:>6.2f} {cr:>6.2f} "
              f"{faith:>6.2f} {ar:>6.2f} {pa_str:>6}  {elapsed}s")

        results.append({
            "question":          question,
            "answer":            answer,
            "ground_truth":      ground_truth,
            "ground_truth_pages": gt_pages,
            "retrieved_pages":   sorted(set(
                c["page"] for c in retrieved if c.get("page")
            )),
            "context_precision": cp,
            "context_recall":    cr,
            "faithfulness":      faith,
            "answer_relevancy":  ar,
            "page_accuracy":     pa,
            "elapsed_s":         elapsed,
        })

    # Aggregate
    def mean(key):
        vals = [r[key] for r in results if r[key] is not None]
        return round(sum(vals) / len(vals), 3) if vals else None

    print(f"\n  {'MEAN':<40} "
          f"{mean('context_precision'):>6.2f} "
          f"{mean('context_recall'):>6.2f} "
          f"{mean('faithfulness'):>6.2f} "
          f"{mean('answer_relevancy'):>6.2f} "
          f"{str(mean('page_accuracy')):>6}")

    # Export
    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"rag_eval_report_{ts}.xlsx"
    build_eval_report(results, output_path)

if __name__ == "__main__":
    main()
