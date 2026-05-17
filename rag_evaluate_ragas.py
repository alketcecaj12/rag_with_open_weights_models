# rag_evaluate_xlsx.py — RAGAS-style evaluation with rich Excel report
#
# Ground truth is loaded from an external JSON file (default: ground_truth.json).
# Never hardcode questions/answers inside this script.
#
# Metrics (all local via Ollama, no external API):
#   1. Context Precision   — retrieval signal-to-noise
#   2. Context Recall      — retrieval completeness
#   3. Faithfulness        — hallucination check
#   4. Answer Relevancy    — end-to-end answer quality
#   5. Page Accuracy       — deterministic page-hit bonus metric
#
# Output: rag_eval_report_<timestamp>.xlsx  (5 sheets)
#   Sheet 1 — Summary        : aggregate score cards + interpretation guide
#   Sheet 2 — Per-Question   : full result table with colour-coded scores
#   Sheet 3 — Category View  : mean scores broken down by question category
#   Sheet 4 — Failure Report : questions with ≥1 metric below 0.40
#   Sheet 5 — Config         : parameter snapshot
#
# Usage:
#   python rag_evaluate_xlsx.py                          # uses ground_truth.json
#   python rag_evaluate_xlsx.py --gt my_questions.json  # custom file
#
# Requirements:
#   pip install chromadb pypdf nltk rank_bm25 openpyxl requests

import re
import sys
import json
import time
import hashlib
import argparse
import requests
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import nltk
import chromadb
from rank_bm25 import BM25Okapi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

nltk.download("punkt",     quiet=True)
nltk.download("punkt_tab", quiet=True)

# ─── CONFIG ──────────────────────────────────────────────────────────────────
OLLAMA_URL        = "http://localhost:11434"
GEN_MODEL         = "phi3:mini"
EMBED_MODEL       = "mxbai-embed-large"
DOCS_DIR          = "./docs"
CHROMA_DIR        = "./chroma_store"
COLLECTION        = "rag_knowledge"
CHUNK_SIZE        = 1500
CHUNK_OVERLAP     = 300
TOP_K             = 7
SCORE_THRESHOLD   = 0.40
RRF_K             = 60
GROUND_TRUTH_FILE = "ground_truth.json"

# ─── GROUND TRUTH LOADER ─────────────────────────────────────────────────────

def load_ground_truth(path: str) -> list[dict]:
    """
    Load evaluation dataset from an external JSON file.

    Expected schema per entry:
      {
        "category":           str,        e.g. "A: Factual/Definition"
        "question":           str,
        "ground_truth":       str,
        "ground_truth_pages": list[int]   empty list for out-of-scope questions
      }
    """
    p = Path(path)
    if not p.exists():
        print(f"  ERROR: Ground truth file not found: {path}")
        sys.exit(1)

    with open(p, "r", encoding="utf-8") as f:
        data = json.load(f)

    required = {"question", "ground_truth", "ground_truth_pages"}
    for i, entry in enumerate(data):
        missing = required - entry.keys()
        if missing:
            print(f"  ERROR: Entry {i} is missing fields: {missing}")
            sys.exit(1)
        if "category" not in entry:
            entry["category"] = "Uncategorised"

    print(f"  Loaded {len(data)} questions from '{path}'")
    return data

# ─── RAG PIPELINE ────────────────────────────────────────────────────────────

def ollama_embed(texts):
    embeddings = []
    for text in texts:
        r = requests.post(
            f"{OLLAMA_URL}/api/embeddings",
            json={"model": EMBED_MODEL, "prompt": text}
        )
        r.raise_for_status()
        embeddings.append(r.json()["embedding"])
    return embeddings

def ollama_generate(prompt):
    r = requests.post(
        f"{OLLAMA_URL}/api/generate",
        json={"model": GEN_MODEL, "prompt": prompt, "stream": False}
    )
    r.raise_for_status()
    return r.json()["response"].strip()

def load_pdf(path):
    from pypdf import PdfReader
    reader = PdfReader(path)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append({"text": text, "page": i + 1})
    return pages

def load_documents(docs_dir):
    docs = []
    for filepath in Path(docs_dir).glob("**/*"):
        if filepath.suffix.lower() == ".pdf":
            pages = load_pdf(str(filepath))
            if pages:
                docs.append({"source": str(filepath), "pages": pages})
        elif filepath.suffix.lower() == ".txt":
            with open(str(filepath), "r", encoding="utf-8") as f:
                text = f.read()
            if text.strip():
                docs.append({
                    "source": str(filepath),
                    "pages":  [{"text": text, "page": None}]
                })
    return docs

def chunk_text(text, chunk_size, overlap):
    from nltk.tokenize import sent_tokenize
    sentences, chunks = sent_tokenize(text), []
    current, current_len = [], 0
    for sent in sentences:
        slen = len(sent)
        if current_len + slen > chunk_size and current:
            chunks.append(" ".join(current))
            kept, kept_len = [], 0
            for s in reversed(current):
                if kept_len + len(s) > overlap:
                    break
                kept.insert(0, s)
                kept_len += len(s)
            current, current_len = kept, kept_len
        current.append(sent)
        current_len += slen
    if current:
        chunks.append(" ".join(current))
    return [c for c in chunks if len(c) > 30]

def chunk_documents(docs):
    all_chunks = []
    for doc in docs:
        for pb in doc["pages"]:
            page_num = pb["page"] or 0
            for i, chunk in enumerate(chunk_text(pb["text"], CHUNK_SIZE, CHUNK_OVERLAP)):
                cid = hashlib.md5(f"{doc['source']}_{page_num}_{i}".encode()).hexdigest()
                all_chunks.append({
                    "id": cid, "text": chunk,
                    "source": doc["source"],
                    "page": page_num, "chunk_index": i,
                })
    return all_chunks

def get_collection():
    client = chromadb.PersistentClient(path=CHROMA_DIR)
    return client.get_or_create_collection(
        name=COLLECTION, metadata={"hnsw:space": "cosine"}
    )

def index_chunks(collection, chunks):
    existing = set(collection.get()["ids"])
    new = [c for c in chunks if c["id"] not in existing]
    if not new:
        print("  ChromaDB: all chunks already indexed.")
        return
    print(f"  Embedding {len(new)} new chunks...")
    for i, chunk in enumerate(new):
        emb = ollama_embed([chunk["text"]])[0]
        collection.add(
            ids=[chunk["id"]], embeddings=[emb],
            documents=[chunk["text"]],
            metadatas=[{
                "source": chunk["source"],
                "page":   chunk["page"],
                "chunk_index": chunk["chunk_index"],
            }]
        )
        if (i + 1) % 100 == 0:
            print(f"    {i+1}/{len(new)}...")
    print("  Done.")

class BM25Index:
    def __init__(self, chunks):
        self.chunks = chunks
        self.bm25   = BM25Okapi([c["text"].lower().split() for c in chunks])

    def search(self, query, top_k):
        scores = self.bm25.get_scores(query.lower().split())
        ranked = sorted(zip(scores, self.chunks),
                        key=lambda x: x[0], reverse=True)[:top_k]
        return [
            {**chunk, "bm25_score": round(float(score), 4)}
            for score, chunk in ranked if score > 0
        ]

def hybrid_retrieve(query, collection, bm25_index):
    qemb = ollama_embed([query])[0]
    res  = collection.query(
        query_embeddings=[qemb], n_results=TOP_K,
        include=["documents", "metadatas", "distances"]
    )
    semantic = []
    for text, meta, dist in zip(
        res["documents"][0], res["metadatas"][0], res["distances"][0]
    ):
        score = round(1 - dist, 3)
        if score >= SCORE_THRESHOLD:
            semantic.append({
                "text": text,
                "source": Path(meta["source"]).name,
                "page":   meta.get("page", 0),
                "sem_score": score,
            })

    bm25_hits = [
        {"text": c["text"], "source": Path(c["source"]).name,
         "page": c["page"], "bm25_score": c["bm25_score"]}
        for c in bm25_index.search(query, TOP_K)
    ]

    rrf_scores, chunk_store = {}, {}
    for rank, chunk in enumerate(semantic):
        key = chunk["text"][:60]
        rrf_scores[key]  = rrf_scores.get(key, 0) + 1 / (rank + RRF_K)
        chunk_store[key] = chunk
    for rank, chunk in enumerate(bm25_hits):
        key = chunk["text"][:60]
        rrf_scores[key]  = rrf_scores.get(key, 0) + 1 / (rank + RRF_K)
        if key not in chunk_store:
            chunk_store[key] = chunk

    results = []
    for key, rrf_score in sorted(rrf_scores.items(),
                                  key=lambda x: x[1], reverse=True)[:TOP_K]:
        chunk = chunk_store[key].copy()
        chunk["rrf_score"] = round(rrf_score, 6)
        results.append(chunk)
    return results

def generate_answer(question, retrieved):
    if not retrieved:
        return "I could not find relevant information in the indexed documents."
    context = "\n\n".join(
        f"[Source: {r['source']}, Page: {r['page']}]\n{r['text']}"
        for r in retrieved
    )
    prompt = f"""Use only the context below to answer the question.
Be concise and factual.
If the answer is not in the context, say "I don't know".
At the end of your answer, cite the source file and page number(s):
  Source: <filename>, Page <number>

Context:
{context}

Question: {question}
Answer:"""
    return ollama_generate(prompt)

# ─── RAGAS METRICS ───────────────────────────────────────────────────────────

def _parse_score(response: str, low=0.0, high=1.0) -> float:
    matches = re.findall(r"\d+(?:\.\d+)?", response)
    for m in matches:
        val = float(m)
        if low <= val <= high:
            return val
    return 0.0

def score_context_precision(question, retrieved, ground_truth) -> float:
    if not retrieved:
        return 0.0
    relevant = 0
    for chunk in retrieved:
        prompt = f"""You are evaluating a RAG retrieval system.

Question: {question}

Retrieved chunk:
{chunk['text'][:600]}

Ground truth answer (for reference):
{ground_truth}

Is this chunk relevant to answering the question?
Answer with only a number: 1 (relevant) or 0 (not relevant)."""
        resp = ollama_generate(prompt)
        relevant += 1 if "1" in resp.split()[:3] else 0
    return round(relevant / len(retrieved), 3)

def score_context_recall(question, retrieved, ground_truth) -> float:
    if not retrieved:
        return 0.0
    context = "\n\n---\n\n".join(c["text"][:400] for c in retrieved)
    prompt = f"""You are evaluating a RAG retrieval system.

Question: {question}

Ground truth answer:
{ground_truth}

Retrieved context (all chunks combined):
{context}

How much of the ground truth answer is covered by the retrieved context?
Score from 0 to 10 where:
  0  = nothing in the ground truth is found in the context
  5  = about half the key facts are present
  10 = all key facts in the ground truth are present in the context

Reply with ONLY a single integer (0-10). No explanation."""
    resp = ollama_generate(prompt)
    raw  = _parse_score(resp, low=0, high=10)
    return round(raw / 10, 3)

def score_faithfulness(question, retrieved, answer) -> float:
    if not retrieved or not answer:
        return 0.0
    context = "\n\n---\n\n".join(c["text"][:400] for c in retrieved)
    prompt = f"""You are evaluating whether an AI answer is grounded in its source context.

Question: {question}

Retrieved context:
{context}

Generated answer:
{answer}

What fraction of the statements in the answer are directly supported by the context?
Score from 0 to 10 where:
  0  = the answer is entirely hallucinated
  5  = about half the claims are supported
  10 = every claim is supported by the context

Reply with ONLY a single integer (0-10). No explanation."""
    resp = ollama_generate(prompt)
    raw  = _parse_score(resp, low=0, high=10)
    return round(raw / 10, 3)

def score_answer_relevancy(question, answer) -> float:
    if not answer:
        return 0.0
    prompt = f"""You are evaluating whether an AI answer is relevant to the question asked.

Question: {question}

Answer:
{answer}

How well does the answer address the question?
Score from 0 to 10 where:
  0  = completely off-topic
  5  = partially relevant
  10 = directly and completely addresses the question

Reply with ONLY a single integer (0-10). No explanation."""
    resp = ollama_generate(prompt)
    raw  = _parse_score(resp, low=0, high=10)
    return round(raw / 10, 3)

def score_page_accuracy(retrieved, ground_truth_pages):
    if not ground_truth_pages:
        return None
    retrieved_pages = set(c["page"] for c in retrieved if c.get("page"))
    hits = len(retrieved_pages & set(ground_truth_pages))
    return round(hits / len(ground_truth_pages), 3)

# ─── EXCEL STYLE HELPERS ─────────────────────────────────────────────────────

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E5D9E"
LIGHT_BLUE = "D9E1F2"
GREEN      = "E2EFDA"
AMBER      = "FFF2CC"
RED        = "FCE4D6"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"
GREEN_FONT = "375623"
AMBER_FONT = "7D6608"
RED_FONT   = "9C0006"

def _fill(c):
    return PatternFill("solid", start_color=c, fgColor=c)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(wrap=True, h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _score_color(score):
    if score is None:
        return GREY, "555555"
    if score >= 0.75:
        return GREEN, GREEN_FONT
    if score >= 0.40:
        return AMBER, AMBER_FONT
    return RED, RED_FONT

def _write_title(ws, row, text, span, bg=DARK_BLUE, size=13):
    ws.merge_cells(span)
    c = ws[span.split(":")[0]]
    c.value     = text
    c.font      = _font(bold=True, color=WHITE, size=size)
    c.fill      = _fill(bg)
    c.alignment = _align(h="center", v="center", wrap=False)
    ws.row_dimensions[row].height = 26

def _score_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col,
                value=f"{value:.2f}" if value is not None else "N/A")
    bg, fg = _score_color(value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=True, color=fg, size=9)
    c.border    = _border()
    c.alignment = _align(h="center", v="center", wrap=False)
    return c

def _mean(results, key):
    vals = [r[key] for r in results if r[key] is not None]
    return round(sum(vals) / len(vals), 3) if vals else None

# ─── METRIC METADATA ─────────────────────────────────────────────────────────

METRIC_LABELS = [
    "Context Precision", "Context Recall",
    "Faithfulness", "Answer Relevancy", "Page Accuracy"
]
METRIC_KEYS = [
    "context_precision", "context_recall",
    "faithfulness", "answer_relevancy", "page_accuracy"
]
METRIC_DESC = [
    "Of the chunks retrieved, what fraction were actually relevant? Low = retriever is noisy.",
    "How much of the correct answer was present in the retrieved chunks? Low = missed the key chunk.",
    "Are the claims in the answer grounded in the retrieved chunks? 1.0 = no hallucination.",
    "Does the answer actually address the question? Catches answers that are faithful but off-topic.",
    "Did the retriever find at least one chunk from the correct page(s)? N/A for out-of-scope questions.",
]

# ─── SHEET 1: SUMMARY ────────────────────────────────────────────────────────

def build_sheet_summary(wb, results):
    ws = wb.active
    ws.title = "Summary"

    _write_title(ws, 1, "RAG Evaluation Report — RAGAS Metrics", "A1:H1")

    ws.merge_cells("A2:H2")
    ws["A2"].value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Model: {GEN_MODEL}  |  Embeddings: {EMBED_MODEL}  |  "
        f"Questions evaluated: {len(results)}"
    )
    ws["A2"].font      = _font(color=WHITE, size=9)
    ws["A2"].fill      = _fill(MID_BLUE)
    ws["A2"].alignment = _align(h="center", wrap=False)
    ws.row_dimensions[2].height = 15

    ws.merge_cells("A4:H4")
    ws["A4"].value     = "What each metric measures"
    ws["A4"].font      = _font(bold=True, color=WHITE, size=10)
    ws["A4"].fill      = _fill(MID_BLUE)
    ws["A4"].alignment = _align(h="center", wrap=False)

    for ri, (name, desc) in enumerate(zip(METRIC_LABELS, METRIC_DESC), start=5):
        nc = ws.cell(row=ri, column=1, value=name)
        dc = ws.cell(row=ri, column=2, value=desc)
        ws.merge_cells(f"B{ri}:H{ri}")
        nc.font = _font(bold=True, size=9)
        dc.font = _font(size=9)
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        for c in (nc, dc):
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = _align(wrap=True, v="center")
        ws.row_dimensions[ri].height = 30

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    ws.row_dimensions[10].height = 10
    _write_title(ws, 11, "Aggregate scores (mean across all questions)",
                 "A11:H11", bg=MID_BLUE, size=10)

    for ci, (label, key) in enumerate(zip(METRIC_LABELS, METRIC_KEYS), start=1):
        avg = _mean(results, key)
        bg, fg = _score_color(avg)

        hc = ws.cell(row=12, column=ci, value=label)
        hc.font      = _font(bold=True, size=9, color="555555")
        hc.fill      = _fill(GREY)
        hc.alignment = _align(h="center", wrap=True)
        hc.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = 20

        vc = ws.cell(row=13, column=ci,
                     value=f"{avg:.2f}" if avg is not None else "N/A")
        vc.font      = _font(bold=True, size=20, color=fg)
        vc.fill      = _fill(bg)
        vc.alignment = _align(h="center", v="center", wrap=False)
        vc.border    = _border()
    ws.row_dimensions[13].height = 40

    ws.row_dimensions[14].height = 10
    _write_title(ws, 15, "Score interpretation guide", "A15:H15",
                 bg=MID_BLUE, size=10)

    guide = [
        ("≥ 0.75",    "Good", GREEN, GREEN_FONT, "System is performing well on this metric."),
        ("0.40–0.74", "Fair", AMBER, AMBER_FONT, "Acceptable but room to improve. Investigate individual failures."),
        ("< 0.40",    "Poor", RED,   RED_FONT,   "Systematic problem. Check chunk size, TOP_K, or embedding model."),
    ]
    for ri, (rng, label, bg, fg, advice) in enumerate(guide, start=16):
        rc = ws.cell(row=ri, column=1, value=rng)
        lc = ws.cell(row=ri, column=2, value=label)
        ac = ws.cell(row=ri, column=3, value=advice)
        ws.merge_cells(f"C{ri}:H{ri}")
        rc.font = _font(bold=True, size=10, color=fg)
        lc.font = _font(bold=True, size=10, color=fg)
        ac.font = _font(size=9)
        for c in (rc, lc, ac):
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = _align(h="center" if c is not ac else "left", wrap=False)
        ws.row_dimensions[ri].height = 18

# ─── SHEET 2: PER-QUESTION ───────────────────────────────────────────────────

def build_sheet_per_question(wb, results):
    ws = wb.create_sheet("Per-Question Results")

    headers    = ["#", "Category", "Question", "Answer",
                  "Ctx\nPrec", "Ctx\nRecall", "Faith-\nfulness", "Ans\nRel",
                  "Page\nAcc", "Pages\nRetrieved", "Pages\nExpected", "Time (s)"]
    col_widths = [4,   20,          38,          52,
                  10,    10,          10,             10,
                  10,    16,              16,              9]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = _font(bold=True, color=WHITE, size=9)
        c.fill      = _fill(DARK_BLUE)
        c.alignment = _align(h="center", v="center", wrap=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    SCORE_COLS = {5, 6, 7, 8, 9}

    for ri, r in enumerate(results, start=2):
        row_bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        row_data = [
            ri - 1,
            r["category"],
            r["question"],
            r["answer"],
            r["context_precision"],
            r["context_recall"],
            r["faithfulness"],
            r["answer_relevancy"],
            r["page_accuracy"],
            ", ".join(f"p.{p}" for p in r["retrieved_pages"]) or "—",
            ", ".join(f"p.{p}" for p in r["ground_truth_pages"]) or "—",
            f"{r['elapsed_s']:.1f}",
        ]
        for ci, val in enumerate(row_data, start=1):
            if ci in SCORE_COLS and isinstance(val, float):
                _score_cell(ws, ri, ci, val)
            elif ci in SCORE_COLS and val is None:
                c = ws.cell(row=ri, column=ci, value="N/A")
                c.fill = _fill(GREY); c.font = _font(size=9, color="555555")
                c.border = _border(); c.alignment = _align(h="center", wrap=False)
            else:
                c = ws.cell(row=ri, column=ci,
                            value=val if val is not None else "N/A")
                c.fill = _fill(row_bg); c.font = _font(size=9)
                c.border = _border(); c.alignment = _align(wrap=True)
        ws.row_dimensions[ri].height = 72

# ─── SHEET 3: CATEGORY VIEW ──────────────────────────────────────────────────

def build_sheet_category(wb, results):
    ws = wb.create_sheet("Category View")
    _write_title(ws, 1, "Mean scores by question category", "A1:G1")

    headers    = ["Category", "N"] + METRIC_LABELS
    col_widths = [28, 5] + [14] * 5
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _font(bold=True, color=WHITE, size=9)
        c.fill      = _fill(MID_BLUE)
        c.alignment = _align(h="center", wrap=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 28

    by_cat = defaultdict(list)
    for r in results:
        by_cat[r["category"]].append(r)

    for ri, (cat, rows) in enumerate(sorted(by_cat.items()), start=3):
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        cc = ws.cell(row=ri, column=1, value=cat)
        cc.fill = _fill(bg); cc.font = _font(size=9, bold=True)
        cc.border = _border(); cc.alignment = _align(wrap=False)
        nc = ws.cell(row=ri, column=2, value=len(rows))
        nc.fill = _fill(bg); nc.font = _font(size=9)
        nc.border = _border(); nc.alignment = _align(h="center", wrap=False)
        for ci, key in enumerate(METRIC_KEYS, start=3):
            _score_cell(ws, ri, ci, _mean(rows, key))
        ws.row_dimensions[ri].height = 18

    tr = len(by_cat) + 3
    for col, val in [(1, "OVERALL"), (2, len(results))]:
        c = ws.cell(row=tr, column=col, value=val)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(DARK_BLUE); c.border = _border()
        c.alignment = _align(wrap=False, h="center" if col == 2 else "left")
    for ci, key in enumerate(METRIC_KEYS, start=3):
        _score_cell(ws, tr, ci, _mean(results, key))
    ws.row_dimensions[tr].height = 20

# ─── SHEET 4: FAILURE REPORT ─────────────────────────────────────────────────

def build_sheet_failures(wb, results):
    ws = wb.create_sheet("Failure Report")
    _write_title(ws, 1,
                 "Failure Report — questions with at least one metric below 0.40",
                 "A1:I1", bg="8B0000")

    failures = [
        r for r in results
        if any(r[k] is not None and r[k] < 0.40 for k in METRIC_KEYS)
    ]

    if not failures:
        ws.merge_cells("A3:I3")
        ws["A3"].value     = "No failures detected — all metrics ≥ 0.40 across all questions."
        ws["A3"].font      = _font(bold=True, size=11, color=GREEN_FONT)
        ws["A3"].fill      = _fill(GREEN)
        ws["A3"].alignment = _align(h="center")
        return

    headers    = ["Category", "Question", "Ctx Prec", "Ctx Recall",
                  "Faithful", "Ans Rel", "Page Acc", "Auto-Diagnosis", "Time (s)"]
    col_widths = [20,          40,          10,          10,
                  10,           10,          10,           45,              9]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _font(bold=True, color=WHITE, size=9)
        c.fill      = _fill("8B0000")
        c.alignment = _align(h="center", wrap=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 28

    for ri, r in enumerate(failures, start=3):
        diag = []
        if r["context_precision"] is not None and r["context_precision"] < 0.40:
            diag.append("Noisy retrieval: irrelevant chunks retrieved.")
        if r["context_recall"] is not None and r["context_recall"] < 0.40:
            diag.append("Missed chunks: try larger CHUNK_SIZE or higher TOP_K.")
        if r["faithfulness"] is not None and r["faithfulness"] < 0.40:
            diag.append("Hallucination: answer not grounded in context.")
        if r["answer_relevancy"] is not None and r["answer_relevancy"] < 0.40:
            diag.append("Off-topic answer: answer did not address the question.")
        if r["page_accuracy"] is not None and r["page_accuracy"] < 0.40:
            diag.append("Wrong page: SCORE_THRESHOLD may be too low.")
        diagnosis = " | ".join(diag) if diag else "Review manually."

        row_data = [
            r["category"], r["question"],
            r["context_precision"], r["context_recall"],
            r["faithfulness"], r["answer_relevancy"], r["page_accuracy"],
            diagnosis, f"{r['elapsed_s']:.1f}",
        ]
        SCORE_COLS_F = {3, 4, 5, 6, 7}
        for ci, val in enumerate(row_data, start=1):
            if ci in SCORE_COLS_F and isinstance(val, float):
                _score_cell(ws, ri, ci, val)
            elif ci in SCORE_COLS_F and val is None:
                c = ws.cell(row=ri, column=ci, value="N/A")
                c.fill = _fill(GREY); c.font = _font(size=9, color="555555")
                c.border = _border(); c.alignment = _align(h="center", wrap=False)
            else:
                c = ws.cell(row=ri, column=ci,
                            value=val if val is not None else "N/A")
                c.fill = _fill(WHITE); c.font = _font(size=9)
                c.border = _border(); c.alignment = _align(wrap=True)
        ws.row_dimensions[ri].height = 72

# ─── SHEET 5: CONFIG ─────────────────────────────────────────────────────────

def build_sheet_config(wb, gt_file, n_questions):
    ws = wb.create_sheet("Config")
    _write_title(ws, 1, "Run Configuration", "A1:B1")

    cfg = [
        ("Parameter",         "Value"),
        ("Ground truth file", gt_file),
        ("Questions",          n_questions),
        ("OLLAMA_URL",         OLLAMA_URL),
        ("GEN_MODEL",          GEN_MODEL),
        ("EMBED_MODEL",        EMBED_MODEL),
        ("DOCS_DIR",           DOCS_DIR),
        ("CHROMA_DIR",         CHROMA_DIR),
        ("COLLECTION",         COLLECTION),
        ("CHUNK_SIZE",         CHUNK_SIZE),
        ("CHUNK_OVERLAP",      CHUNK_OVERLAP),
        ("TOP_K",              TOP_K),
        ("SCORE_THRESHOLD",    SCORE_THRESHOLD),
        ("RRF_K",              RRF_K),
        ("Run timestamp",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for ri, (k, v) in enumerate(cfg, start=2):
        ck = ws.cell(row=ri, column=1, value=k)
        cv = ws.cell(row=ri, column=2, value=v)
        if ri == 2:
            for c in (ck, cv):
                c.font = _font(bold=True, color=WHITE)
                c.fill = _fill(MID_BLUE)
        else:
            bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
            for c in (ck, cv):
                c.fill = _fill(bg); c.font = _font(size=10)
        for c in (ck, cv):
            c.border    = _border()
            c.alignment = _align(wrap=False)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45

# ─── REPORT ENTRY POINT ──────────────────────────────────────────────────────

def build_eval_report(results, output_path, gt_file):
    wb = Workbook()
    build_sheet_summary(wb, results)
    build_sheet_per_question(wb, results)
    build_sheet_category(wb, results)
    build_sheet_failures(wb, results)
    build_sheet_config(wb, gt_file, len(results))
    wb.save(output_path)
    print(f"\n  Report saved: {output_path}")

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="RAG evaluation with Excel report")
    parser.add_argument("--gt", default=GROUND_TRUTH_FILE,
                        help=f"Path to ground truth JSON (default: {GROUND_TRUTH_FILE})")
    args = parser.parse_args()

    print("\n=== RAG Evaluation — RAGAS Metrics ===\n")

    print(f"[0/3] Loading ground truth from '{args.gt}'...")
    ground_truth = load_ground_truth(args.gt)

    print(f"\n[1/3] Loading documents from '{DOCS_DIR}'...")
    docs = load_documents(DOCS_DIR)
    if not docs:
        print("  No documents found. Add PDFs to ./docs/ and rerun.")
        return
    chunks = chunk_documents(docs)
    print(f"  {len(chunks)} chunks ready.")

    print(f"\n[2/3] Setting up indexes...")
    collection = get_collection()
    index_chunks(collection, chunks)
    bm25_index = BM25Index(chunks)

    print(f"\n[3/3] Evaluating {len(ground_truth)} questions...\n")
    print(f"  {'Question':<40} {'Prec':>6} {'Rec':>6} {'Faith':>6} {'Rel':>6} {'Page':>6}  Time")
    print(f"  {'-'*40} {'-'*6} {'-'*6} {'-'*6} {'-'*6} {'-'*6}  ----")

    results = []
    for entry in ground_truth:
        question  = entry["question"]
        gt_answer = entry["ground_truth"]
        gt_pages  = entry["ground_truth_pages"]
        category  = entry["category"]

        t0        = time.time()
        retrieved = hybrid_retrieve(question, collection, bm25_index)
        answer    = generate_answer(question, retrieved)

        cp    = score_context_precision(question, retrieved, gt_answer)
        cr    = score_context_recall(question, retrieved, gt_answer)
        faith = score_faithfulness(question, retrieved, answer)
        ar    = score_answer_relevancy(question, answer)
        pa    = score_page_accuracy(retrieved, gt_pages)

        elapsed = round(time.time() - t0, 1)
        pa_str  = f"{pa:.2f}" if pa is not None else " N/A"
        print(f"  {question[:40]:<40} {cp:>6.2f} {cr:>6.2f} "
              f"{faith:>6.2f} {ar:>6.2f} {pa_str:>6}  {elapsed}s")

        results.append({
            "category":           category,
            "question":           question,
            "answer":             answer,
            "ground_truth":       gt_answer,
            "ground_truth_pages": gt_pages,
            "retrieved_pages":    sorted(set(
                c["page"] for c in retrieved if c.get("page")
            )),
            "context_precision":  cp,
            "context_recall":     cr,
            "faithfulness":       faith,
            "answer_relevancy":   ar,
            "page_accuracy":      pa,
            "elapsed_s":          elapsed,
        })

    print(f"\n  {'MEAN':<40} "
          f"{_mean(results, 'context_precision'):>6.2f} "
          f"{_mean(results, 'context_recall'):>6.2f} "
          f"{_mean(results, 'faithfulness'):>6.2f} "
          f"{_mean(results, 'answer_relevancy'):>6.2f} "
          f"{str(_mean(results, 'page_accuracy')):>6}")

    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"rag_eval_report_{ts}.xlsx"
    build_eval_report(results, output_path, args.gt)


if __name__ == "__main__":
    main()
